#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抖音视频批量文案处理器
配合 AnyToCopy 使用，批量处理提取的文案并生成汇总 Excel

使用方法:
1. 将 AnyToCopy 提取的文案保存为 JSON 文件
2. 运行此脚本批量处理
3. 生成汇总 Excel 文件
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 配置
# 获取脚本所在目录的父目录（项目根目录）
import sys
from pathlib import Path
PROJECT_ROOT = Path(__file__).parent.parent
INPUT_DIR = PROJECT_ROOT / "input" / "douyin_texts"  # 输入目录
OUTPUT_DIR = PROJECT_ROOT / "output"  # 输出目录

# 确保目录存在
INPUT_DIR = str(INPUT_DIR)
OUTPUT_DIR = str(OUTPUT_DIR)
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_video_texts():
    """从输入目录加载所有视频文案 JSON 文件"""
    videos = []
    
    # 查找所有 JSON 文件
    json_files = [f for f in os.listdir(INPUT_DIR) if f.endswith('.json')]
    
    if not json_files:
        print(f"⚠️  在 {INPUT_DIR} 目录下未找到 JSON 文件")
        print(f"💡 请将 AnyToCopy 提取的文案保存为 JSON 文件到此目录")
        return videos
    
    print(f"📂 找到 {len(json_files)} 个文案文件")
    
    for json_file in sorted(json_files):
        try:
            file_path = os.path.join(INPUT_DIR, json_file)
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                videos.append(data)
                print(f"  ✓ 已加载：{json_file}")
        except Exception as e:
            print(f"  ✗ 加载失败 {json_file}: {e}")
    
    return videos


def create_summary_excel(videos, output_file):
    """创建汇总 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文案汇总"
    
    # 样式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    normal_font = Font(size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    link_font = Font(color="0066CC", underline="single")
    
    # 表头
    headers = ["序号", "视频标题", "作者", "时长", "发布时间", "完整文案", "视频链接"]
    for col, header in enumerate(headers, start=1):
        cell = ws[f'{get_column_letter(col)}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据行
    for idx, video in enumerate(videos, start=2):
        ws[f'A{idx}'] = idx - 1
        ws[f'B{idx}'] = video.get('标题', '')
        ws[f'C{idx}'] = video.get('作者', '')
        ws[f'D{idx}'] = video.get('时长', '')
        ws[f'E{idx}'] = video.get('发布时间', '')
        ws[f'F{idx}'] = video.get('完整文案', video.get('简介文案', ''))
        ws[f'G{idx}'] = video.get('视频链接', '')
        
        # 格式设置
        for col in range(1, 8):
            cell = ws[f'{get_column_letter(col)}{idx}']
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 链接样式
        if video.get('视频链接'):
            ws[f'G{idx}'].font = link_font
    
    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 80
    ws.column_dimensions['G'].width = 50
    
    # 添加统计信息
    last_row = len(videos) + 2
    ws[f'A{last_row}'] = f"总计：{len(videos)} 个视频"
    ws[f'A{last_row}'].font = Font(bold=True)
    ws[f'B{last_row}'] = f"处理时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    wb.save(output_file)
    return output_file


def create_detail_excel(videos, output_file):
    """创建详细文案 Excel（每个视频一个 Sheet）"""
    wb = Workbook()
    
    # 删除默认 Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 样式
    title_font = Font(bold=True, size=14)
    info_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    
    for idx, video in enumerate(videos, start=1):
        sheet_name = f"视频{idx}"[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        
        # 标题
        ws['A1'] = "视频文案提取结果"
        ws['A1'].font = title_font
        
        # 基本信息
        row = 3
        info_fields = [
            ('视频链接', video.get('视频链接', '')),
            ('视频标题', video.get('标题', '')),
            ('作者', video.get('作者', '')),
            ('视频时长', video.get('时长', '')),
            ('发布时间', video.get('发布时间', '')),
            ('简介文案', video.get('简介文案', '')),
        ]
        
        for label, value in info_fields:
            if value:
                ws[f'A{row}'] = f"{label}:"
                ws[f'A{row}'].font = info_font
                ws[f'B{row}'] = value
                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                row += 1
        
        # 完整文案
        row += 1
        ws[f'A{row}'] = "完整文案:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        full_text = video.get('完整文案', video.get('简介文案', ''))
        ws[f'A{row}'] = full_text
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # 列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80
    
    wb.save(output_file)
    return output_file


def create_stats_excel(videos, output_file):
    """创建统计信息 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "视频统计"
    
    # 样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 表头
    headers = ["序号", "视频标题", "作者", "字数", "时长"]
    for col, header in enumerate(headers, start=1):
        cell = ws[f'{get_column_letter(col)}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 数据
    total_chars = 0
    for idx, video in enumerate(videos, start=2):
        text = video.get('完整文案', video.get('简介文案', ''))
        char_count = len(text)
        total_chars += char_count
        
        ws[f'A{idx}'] = idx - 1
        ws[f'B{idx}'] = video.get('标题', '')[:40]
        ws[f'C{idx}'] = video.get('作者', '')
        ws[f'D{idx}'] = char_count
        ws[f'E{idx}'] = video.get('时长', '')
    
    # 统计行
    last_row = len(videos) + 2
    ws[f'A{last_row}'] = "总计"
    ws[f'A{last_row}'].font = Font(bold=True)
    ws[f'D{last_row}'] = total_chars
    ws[f'D{last_row}'].font = Font(bold=True)
    
    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    
    wb.save(output_file)
    return output_file


def main():
    """主函数"""
    print("=" * 60)
    print("🎬 抖音视频批量文案处理器")
    print("=" * 60)
    
    # 加载文案
    videos = load_video_texts()
    
    if not videos:
        print("\n❌ 没有可处理的视频文案")
        print("\n📝 使用步骤:")
        print("1. 将 AnyToCopy 提取的文案保存为 JSON 文件")
        print(f"2. 放入目录：{INPUT_DIR}")
        print("3. 重新运行此脚本")
        return
    
    # 生成时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 创建三个 Excel 文件
    print(f"\n📊 正在生成 Excel 文件...")
    
    # 1. 汇总文件
    summary_file = os.path.join(OUTPUT_DIR, f"抖音文案汇总_{len(videos)}个视频_{timestamp}.xlsx")
    create_summary_excel(videos, summary_file)
    print(f"  ✓ 汇总文件：{summary_file}")
    
    # 2. 详细文件
    detail_file = os.path.join(OUTPUT_DIR, f"抖音文案详细_{len(videos)}个视频_{timestamp}.xlsx")
    create_detail_excel(videos, detail_file)
    print(f"  ✓ 详细文件：{detail_file}")
    
    # 3. 统计文件
    stats_file = os.path.join(OUTPUT_DIR, f"抖音文案统计_{len(videos)}个视频_{timestamp}.xlsx")
    create_stats_excel(videos, stats_file)
    print(f"  ✓ 统计文件：{stats_file}")
    
    # 统计信息
    print(f"\n📝 处理完成!")
    print(f"   视频数量：{len(videos)}")
    total_chars = sum(len(v.get('完整文案', v.get('简介文案', ''))) for v in videos)
    print(f"   总字数：{total_chars:,}")
    print(f"   输出目录：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
